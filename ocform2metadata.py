import pandas as pd
import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Alignment

import re
import os
import sys
import time

import argparse

parser = argparse.ArgumentParser()

parser.add_argument(
    "form_def", help="Name of OC Form Definition File (Excel file)")
args = parser.parse_args()

full_path = args.form_def

if not os.path.exists(full_path):
    sys.exit(f"'File not found - {full_path}")

# Read in each of the worksheets in the Form Definition spreadsheet
df_settings = pd.read_excel(full_path, 'settings', keep_default_na=False, dtype=str)
df_choices = pd.read_excel(full_path, 'choices', keep_default_na=False, dtype=str)
df_survey = pd.read_excel(full_path, 'survey', keep_default_na=False, dtype=str)

header_row = ('Code', 'Type', 'Description', 'Length', 'Format')
# Create empty datafram with column headings
df_metadata = pd.DataFrame(columns=header_row)

# Insert the Form Name as the first row
form_title = str(df_settings['form_title'][0])
metadata_row = {
    'Code': '',
    'Type': 'Form',
    'Description': form_title,
    'Length': '',
    'Format': ''
}
df_metadata = df_metadata.append(metadata_row, ignore_index=True)

# Go through each row of the Survey worksheet and insert an appropriate row in the Metadata
for row in df_survey.itertuples():

    quest_type = str(row.type)

    # If a Select question change to Category
    if quest_type.startswith('select_'):
        quest_type = 'Category'

    # Create the new row
    new_row = {
        'Code': row.name,
        'Type': quest_type.capitalize(),
        'Description': row.label,
        'Length': '',
        'Format': ''
    }
    # Append the new row
    df_metadata = df_metadata.append(new_row, ignore_index=True)

    # Repeating?
    if quest_type.endswith('group') or quest_type.endswith('repeat'):
        repeat_group = 'Yes' if quest_type.endswith('repeat') else 'No'
        # Create the new row
        new_row = {
            'Code': '',
            'Type': 'Repeating:',
            'Description': repeat_group,
            'Length': '',
            'Format': ''
        }
        # Append the new row
        df_metadata = df_metadata.append(new_row, ignore_index=True)

    # If select_ type question, append list choices
    if str(row.type).startswith('select_'):
        listname = str(row.type).split()[1]
        list_choices = df_choices[df_choices['list_name'] == listname]
        for choice in list_choices.itertuples():
            if pd.isna(choice.label) : print(f"{choice}")
            # Create the new row
            new_row = {
                'Code': '',
                'Type': choice.name,
                'Description': choice.label,
                'Length': '',
                'Format': ''
            }
            # Append the new row
            df_metadata = df_metadata.append(new_row, ignore_index=True)

    # Hint
    if row.hint != '':
        new_row = {
            'Code': '',
            'Type': 'Note:',
            'Description': f"Hint: {row.hint}",
            'Length': '',
            'Format': ''
        }
        df_metadata = df_metadata.append(new_row, ignore_index=True)

    # Required
    if row.required != '' and row.required != 'yes':
        new_row = {
            'Code': '',
            'Type': 'REQUIRED-IF:',
            'Description': row.required,
            'Length': '',
            'Format': ''
        }
        df_metadata = df_metadata.append(new_row, ignore_index=True)

    # Relevant
    if row.relevant != '':
        new_row = {
            'Code': '',
            'Type': 'COLLECT-IF',
            'Description': row.relevant,
            'Length': '',
            'Format': ''
        }
        df_metadata = df_metadata.append(new_row, ignore_index=True)

    # Constraint
    if row.constraint != '':
        new_row = {
            'Code': '',
            'Type': 'WARN-IF:',
            'Description': row.constraint_message,
            'Length': '',
            'Format': ''
        }
        df_metadata = df_metadata.append(new_row, ignore_index=True)

    # Calculation
    if row.calculation != '':
        new_row = {
            'Code': '',
            'Type': 'CALCULATE:',
            'Description': row.calculation,
            'Length': '',
            'Format': ''
        }
        df_metadata = df_metadata.append(new_row, ignore_index=True)

# Convert the Metadata dataframe into an Excel object
# https://openpyxl.readthedocs.io/en/stable/pandas.html#:~:text=Working%20with%20Pandas%20Dataframes%20%C2%B6%20The%20openpyxl.utils.dataframe.dataframe_to_rows%20%28%29,wb.active%20for%20r%20in%20dataframe_to_rows%28df%2C%20index%3DTrue%2C%20header%3DTrue%29%3A%20ws.append%28r%29
wb = op.Workbook()
ws = wb.active

# Worksheet title
ws.title = re.sub(r'\W+', ' ',form_title)

# Read each row in the dataframe and add it to the worksheet
for r in dataframe_to_rows(df_metadata, index=False, header=True):
    ws.append(r)

# Format the Header Row
for c in ws['A1:E1'][0]:
    c.font = Font(color='FFFFFF')  # White
    c.fill = PatternFill('solid', fgColor='808080')  # Grey

# Format the Form Name row
for c in ws['A2:E2'][0]:
    c.font = Font(color='000000')  # White
    c.fill = PatternFill('solid', fgColor='90EE90')  # Green

# Format the colours
for r in ws:
    quest_type = str(r[1].value)
    if quest_type.lower().endswith('group') or quest_type.lower().endswith('repeat'):
        for c in r:
            c.fill = PatternFill('solid', fgColor='FFD800')
        continue

    if quest_type == 'Note:' or quest_type == 'Note':
        for c in r:
            c.font = Font(color='FF0000')
        continue

    if quest_type == 'Calculate':
        for c in r:
            c.fill = PatternFill('solid', fgColor='FCD5B4')
            c.alignment = Alignment(vertical='top', wrap_text=True)
        continue

    if quest_type == 'CALCULATE:':
        for c in r:
            c.alignment = Alignment(vertical='top', wrap_text=True)
        continue

    if quest_type in ['Category', 'Text', 'Integer', 'Decimal', 'Date']:
        for c in r:
            c.fill = PatternFill('solid', fgColor='AED8E6')
            r[1].fill = PatternFill('solid', fgColor='AED8E6')
        continue

# Save the Excel object as an Excel file
file_path = os.path.dirname(full_path)
file_name = os.path.basename(full_path)
ts = time.time()
dest = f"{file_path}\MD-{file_name}.xlsx"
wb.save(dest)
